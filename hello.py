import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from openpyxl import Workbook
from openpyxl.worksheet.merge import MergeCell, MergedCell
wb = Workbook()
from copy import copy
from openpyxl.worksheet.copier import WorksheetCopy
import re


from excel_documen_settings import TORG_12_TABLE_CELLS
from excel_documen_settings import TORG_12_TABLE_RESULT_CELLS


from excel_documen_settings import TORG_12_START_TABLE_HEAD
from excel_documen_settings import TORG_12_START_PAGE_PRINT_AREA
from excel_documen_settings import TORG_12_END_PAGE_PRINT_AREA
from excel_documen_settings import TORG_12_END_PRODUCT_TABLE


def fill_cells(sheet):
    sheet['A7'] = 'Поставщикт'
    sheet['I14'] = 'Закупщик'
    sheet['I14'] = 'Поставщикт'
    sheet['I16'] = 'Закупщик'
    sheet['I18'] = 'Основание'
    sheet['AX26'] = 'Номер документа'
    sheet['BI26'] = 'Дата создания тн'
    sheet['K33'] = 7
    # sheet['AR31'] = 'Итого количество'
    # sheet['AR32'] = 'Всего количество'
    # sheet['BQ32'] = 'Сумма без ндс'


def style_cell(sheet, cell):
    sheet[cell].border = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    sheet[cell].alignment = Alignment(
                    horizontal='general',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
    sheet[cell].font = Font(
        name='Calibri',
        size=11,
                 bold=False,
                 italic=False,
                 vertAlign=None,
                 underline='none',
                 strike=False,
                 color='FF000000')


# def delete_merged_cell(row):
#     count = len(sheet.merged_cells.ranges)
#     try:
#         for i in range(count):
#             if re.findall(r'%s' % str(row), sheet.merged_cells.ranges[i].__str__()):
#                 end_list = len(sheet.merged_cells.ranges) - 1
#                 el = sheet.merged_cells.ranges[i]
#                 sheet.merged_cells.ranges[i] = sheet.merged_cells.ranges[end_list]
#                 sheet.merged_cells.ranges[end_list] = el
#                 sheet.merged_cells.ranges.remove(sheet.merged_cells.ranges[end_list])
#     except IndexError:
#         pass


def copy_merged_cells(original_sheet, distenetion_sheet, start_row):
    col = 1
    row_f = 1
    save_srt_row = start_row -1
    for row in original_sheet.rows:
        rd_f = original_sheet.row_dimensions[row_f]
        rd_h = sheet_head.row_dimensions[start_row]
        rd_h.height = rd_f.height
        row_f += 1
        start_row += 1
    for row in original_sheet.rows:
        for cell in row:
            for c in cell.parent.merged_cells.ranges:
                if re.findall(r'^(?:^|\W)%s(?:$|\W)' % str(cell.coordinate), c.coord):
                    distenetion_sheet.merge_cells(
                        start_column=c.min_col,
                        start_row=c.min_row + save_srt_row,
                        end_column=c.max_col,
                        end_row=c.max_row + save_srt_row,
                    )
                    addres = str(c.min_row + save_srt_row)
                    my_cell = ''.join(x for x in cell.coordinate if x.isalpha()) + addres
                    merged_my_cell = distenetion_sheet.merged_cells.ranges[len(distenetion_sheet.merged_cells.ranges) - 1].__str__()
                    for col_1 in distenetion_sheet[merged_my_cell]:
                        for col_2 in col_1:
                            col_2._style = copy(cell._style)
                    distenetion_sheet[my_cell].value = copy(cell.value)
            col += 1
        col = 1
        start_row += 1


def copy_simple_cells(original_sheet, distenetion_sheet, start_row):
    col =1
    for row in original_sheet.rows:
        for cell in row:
            new_cell = distenetion_sheet.cell(
                row=start_row,
                column=col,
                value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
            col += 1
        col = 1
        start_row += 1


def fill_product_table(sheet, sheet_table, start_table_head, end_page, row_fill_remained=70):
    # TODO do not forget about change to real data
    copy_merged_cells(sheet_table, sheet, start_table_head)
    global TORG_12_END_PRODUCT_TABLE
    #for example
    start_prod = start_table_head + 3
    for row in range(start_prod, row_fill_remained):
        sheet.insert_rows(row)
        rd = sheet.row_dimensions[row]
        rd.height = 12
        for key, val in TORG_12_TABLE_CELLS.items():
            dict_cells = TORG_12_TABLE_CELLS.get(key)
            for simple_cell, merge_cell in dict_cells.items():
                merg_cell = merge_cell[0]+str(row) + ':' + merge_cell[1]+str(row)
                cell = simple_cell + str(row)
                style_cell(sheet, cell)
                sheet.merge_cells(merg_cell)
                sheet[cell].value = 'val' + str(row)
        end_page += 1
        TORG_12_END_PRODUCT_TABLE =+ 1
        if end_page >= TORG_12_END_PAGE_PRINT_AREA:
            row_fill_remained = row_fill_remained - end_page
            start_table_head += end_page
            return fill_product_table(sheet, sheet_table, start_table_head, 1, row_fill_remained)


def add_tables(main_sheet, sheet_table, start_table_head, end_page):
    end_page += start_table_head
    fill_product_table(main_sheet, sheet_table, start_table_head, end_page)


def add_footer(sheet_footer, main_sheet, footer_start_row):
    copy_merged_cells(sheet_footer, main_sheet, footer_start_row)
    copy_simple_cells(sheet_footer, main_sheet, footer_start_row)


if __name__ == '__main__':
    work_book = openpyxl.load_workbook(
        'torg-12.xlsm'
    )

    sheet_head = work_book['Head copy']
    sheet_table = work_book['table']
    sheet_footer = work_book['footer']
    start_page = TORG_12_START_PAGE_PRINT_AREA
    start_table_head = TORG_12_START_TABLE_HEAD
    # copy_merged_cells(sheet_table, sheet_head, TORG_12_START_TABLE_HEAD)
    # fill_product_table(sheet_head)
    add_tables(sheet_head, sheet_table, start_table_head, 1)
    add_footer(sheet_footer, sheet_head, 100)
    # copy_filled_sheet(sheet_table, sheet_head, start_sheet)
    # sheet_head.print_area = "A5:CF95"


    # copy_filled_sheet(sheet_footer, sheet_head, 52)
    # fill_footer(sheet_footer, 51)

    work_book.save("test_result_wb.xlsx")

